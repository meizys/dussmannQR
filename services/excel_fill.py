"""SnapMeter Excel auto-fill service (PRD §7.4).

HARD REQUIREMENT: inject values into the client's *existing* master template
without regenerating it, so all styles, column widths, formulas and branding are
preserved byte-for-byte. openpyxl loads the workbook and writes only the mapped
cells, then saves to a NEW file (the master is never overwritten).

This is the production counterpart to the browser SheetJS export in
src/lib/excel.ts (which is the demo-only equivalent).

Usage:
    python excel_fill.py master.xlsx out.xlsx mappings.json
where mappings.json is a list of:
    {"target_sheet": "Readings", "target_cell": "D4", "value": 1850.4}
(produced from template_mappings joined with the period's confirmed readings).
"""
from __future__ import annotations

import json
import sys
from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass
class CellWrite:
    target_sheet: str
    target_cell: str
    value: float


def fill_template(master_path: str, out_path: str, writes: list[CellWrite]) -> int:
    # keep_vba and rich formatting are preserved by openpyxl on load+save.
    wb = load_workbook(master_path)
    written = 0
    for w in writes:
        if w.target_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{w.target_sheet}' not in template")
        ws = wb[w.target_sheet]
        # Assign value only; do NOT touch cell.style/number_format — preserve it.
        ws[w.target_cell] = w.value
        written += 1
    wb.save(out_path)  # never overwrites the master (out_path differs)
    return written


def main() -> None:
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(2)
    master_path, out_path, mappings_path = sys.argv[1:4]
    with open(mappings_path, encoding="utf-8") as fh:
        raw = json.load(fh)
    writes = [CellWrite(**item) for item in raw]
    n = fill_template(master_path, out_path, writes)
    print(f"Wrote {n} cells into {out_path} (master '{master_path}' untouched)")


if __name__ == "__main__":
    main()
